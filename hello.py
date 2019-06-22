from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
import tkinter.ttk as ttk
import os
import openpyxl
import time
import threading
path_to_first_file = ""
path_to_second_file = ""
EXIT = False
WORK = False


def choose_first_excel_file():
    global path_to_first_file
    path_to_first_file = filedialog.askopenfilename(filetypes=(
        ("Excel File", "*.xlsx"), ("Excel File", "*.xlsm"), ("Excel File", "*.xltx"), ("Excel File", "*.xltm")))
    if path_to_first_file != "":
        path_to_first_file = os.path.abspath(path_to_first_file)
        first_file_text.configure(state=NORMAL)
        first_file_text.delete(1.0, END)
        first_file_text.insert(END, path_to_first_file)
        first_file_text.configure(state=DISABLED)


def choose_second_excel_file():
    global path_to_second_file
    path_to_second_file = filedialog.askopenfilename(filetypes=(
        ("Excel File", "*.xlsx"), ("Excel File", "*.xlsm"), ("Excel File", "*.xltx"), ("Excel File", "*.xltm")))
    if path_to_second_file != "":
        path_to_second_file = os.path.abspath(path_to_second_file)
        second_file_text.configure(state=NORMAL)
        second_file_text.delete(1.0, END)
        second_file_text.insert(END, path_to_second_file)
        second_file_text.configure(state=DISABLED)


def ok_click():
    global WORK
    if path_to_first_file != "" and path_to_second_file != "":
        WORK = True
        ProgressBar.place(x=60, y=150, width=400)
    else:
        messagebox.showerror("Incomplete Form", "Пожалуйста заполните все поля")

def read_excel_file(path, pb_counter):
    stream = openpyxl.load_workbook(path, read_only=True)
    string = stream[stream.sheetnames[0]]["A1"].value
    if string == "ВЕДОМОСТЬ":
        stream.save(path)
        return read_excel_vedom(path, pb_counter)
    else:
        return read_excel_our_file(path, pb_counter)

def read_excel_our_file(path, pb_counter):
    global ProgressBar
    stream = openpyxl.load_workbook(path, read_only=True)
    sheet = stream[stream.sheetnames[0]]
    mapa = {}
    iter = 2
    path_char = 'A'
    hash_char = 'B'
    while sheet[path_char + str(iter)].value:
        mapa[sheet[path_char + str(iter)].value] = str(sheet[hash_char + str(iter)].value).lower()
        iter += 1
        ProgressBar['value'] += 80/pb_counter
    return mapa


def read_excel_vedom(path, pb_counter):
    global ProgressBar
    stream = openpyxl.load_workbook(path, read_only=True)
    sheet = stream[stream.sheetnames[0]]
    mapa = {}
    iter = 7
    path_char = 'E'
    hash_char = 'F'
    while sheet[path_char + str(iter)].value:
        mapa[sheet[path_char + str(iter)].value] = str(sheet[hash_char + str(iter)].value).lower()
        iter += 1
        ProgressBar['value'] += 80/pb_counter
    return mapa


def check():
    global WORK
    global ProgressBar
    global EXIT
    while 1:
        time.sleep(0.5)
        if WORK:
            time_start = time.time()
            first_excel_stream = openpyxl.load_workbook(path_to_first_file)
            second_excel_stream = openpyxl.load_workbook(path_to_second_file)
            first_excel_stream.save(path_to_first_file)
            second_excel_stream.save(path_to_second_file)
            pb_counter = first_excel_stream[first_excel_stream.sheetnames[0]].max_row + second_excel_stream[second_excel_stream.sheetnames[0]].max_row
            first_file = read_excel_file(path_to_first_file, pb_counter)
            second_file = read_excel_file(path_to_second_file, pb_counter)
            differences = []
            pb_counter = len(first_file.keys())
            for i in first_file.keys():
                ProgressBar['value'] += 20/pb_counter
                try:
                    if first_file[i] != second_file[i]:
                        differences.append(i)
                except KeyError:
                    continue
            if len(differences) == 0:
                messagebox.showinfo("OK", "Работа Выполнена\nВсе МД5 суммы одинаковы\nБыло затрачено {} сек".format(
                    round(time.time() - time_start, 2)))
            else:
                file_stream_out = open("Различные.txt", 'w', encoding='utf-16')
                file_stream_out.write(u'\n'.join(differences))
                file_stream_out.close()
                messagebox.showinfo("OK",
                                    """Работа Выполнена\n
                                    Названия несовпадающих файлов записаны в файл {}\Несовпадающие.txt\n
                                    Количество различных:{}\n
                                    Было затрачено {} сек""".format(
                                        os.path.abspath(""), len(differences), round(time.time() - time_start, 2)))
            ProgressBar.place(x=10, y=20, width=0)
            ProgressBar['value'] = 0
            WORK = False
        if EXIT:
            exit()


thread1 = threading.Thread(target=check)
thread1.start()
window = Tk()
window.title("Checker v:0.0.1")
window.geometry("500x200")
window.maxsize(500, 200)
window.minsize(500, 200)
first_button = Button(window, text="Выбрать Файл", command=choose_first_excel_file)
second_button = Button(window, text="Выбрать Файл", command=choose_second_excel_file)
first_file_text = Text(window, height=1, width=45)
second_file_text = Text(window, height=1, width=45)
ok_button = Button(window, text="Проверить", command=ok_click)
first_file_text_scrollbar = Scrollbar(window, orient=HORIZONTAL, command=first_file_text.yview)
second_file_text_scrollbar = Scrollbar(window, orient=HORIZONTAL, command=second_file_text.yview)
first_file_text.configure(yscrollcommand=first_file_text_scrollbar.set, state=DISABLED)
second_file_text.configure(yscrollcommand=second_file_text_scrollbar.set, state=DISABLED)
ProgressBar = ttk.Progressbar(window, mode="determinate")
first_button.place(x=10, y=20)
second_button.place(x=10, y=70)
first_file_text.place(x=105, y=23)
second_file_text.place(x=105, y=73)
first_file_text_scrollbar.place(x=105, y=45, width=365, height=10)
second_file_text_scrollbar.place(x=105, y=95, width=365, height=10)
ok_button.place(x=210, y=120, width=90)
window.mainloop()
EXIT = True
